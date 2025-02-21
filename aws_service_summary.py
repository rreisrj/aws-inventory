# aws_service_summary.py
import concurrent.futures
import configparser
import os
import sys
from datetime import datetime
from pathlib import Path
from threading import Lock

import boto3
import botocore
import openpyxl
import pandas as pd
from botocore.exceptions import ClientError
from openpyxl.styles import Font, PatternFill

from aws_components import initialize_all_components


class AWSResourceInventory:
    def __init__(self):
        self.session = None
        self.account_id = None
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.resources = {}
        self.resources_lock = Lock()  # Add lock for thread safety
        self.instance_counts = {}
        self.rds_instance_counts = {}
        self.regions = []
        self.available_regions = [
            "us-east-1",
            "us-east-2",
            "sa-east-1",
            "us-west-1",
            "us-west-2",
        ]
        # Updated to include EFS and KMS
        self.supported_services = [
            "APIGateway",
            "AutoScaling",
            "DynamoDB",
            "CloudFront",
            "EC2",
            "ECR",
            "ECS",
            "EFS",
            "EKS",
            "ELB",
            "Gateway",
            "KMS",
            "Lambda",
            "RDS",
            "Route53",
            "S3",
            "SNS",
            "SQS",
            "TargetGroup",
            "UnattachedEBS",
            "UnattachedSG",
            "UnattachedEIP",
            "UnattachedENI",
            "VPC",
        ]
        self.components = {}

    def select_regions(self):
        """Let user select which regions to scan"""
        print("\nAvailable regions:")
        for i, region in enumerate(self.available_regions, 1):
            print(f"{i}. {region}")

        print("\nSelect regions (comma-separated numbers, or 'all' for all regions)")
        while True:
            selection = input("\nEnter selection: ").strip().lower()
            if selection == "all":
                self.regions = self.available_regions
                break
            try:
                indices = [int(x.strip()) - 1 for x in selection.split(",")]
                if all(0 <= i < len(self.available_regions) for i in indices):
                    self.regions = [self.available_regions[i] for i in indices]
                    break
            except ValueError:
                print(
                    "Invalid input. Please enter numbers separated by commas or 'all'"
                )

    def get_aws_profile(self):
        """List and select AWS profile"""
        profiles = self._list_profiles()
        if not profiles:
            print("No AWS profiles found.")
            sys.exit(1)

        print("\nAvailable AWS profiles:")
        for i, profile in enumerate(profiles, 1):
            print(f"{i}. {profile}")

        while True:
            try:
                index = int(input("\nSelect profile number: ")) - 1
                if 0 <= index < len(profiles):
                    return profiles[index]
            except ValueError:
                pass
            print("Invalid selection. Please try again.")

    def _list_profiles(self):
        """Get list of AWS profiles"""
        profiles = set()
        aws_folder = Path.home() / ".aws"

        # Check credentials file
        credentials_path = aws_folder / "credentials"
        if credentials_path.exists():
            config = configparser.ConfigParser()
            config.read(credentials_path)
            profiles.update(config.sections())

        # Check config file
        config_path = aws_folder / "config"
        if config_path.exists():
            config = configparser.ConfigParser()
            config.read(config_path)
            for section in config.sections():
                # Remove 'profile ' prefix if present
                profile_name = section.replace("profile ", "")
                profiles.add(profile_name)

        if not profiles:
            print("\nNo AWS profiles found in ~/.aws/credentials or ~/.aws/config")
            print("Please configure your AWS credentials first using 'aws configure'")
            sys.exit(1)

        return sorted(profiles)

    def initialize_components(self):
        """Initialize all service components"""
        self.components = initialize_all_components(self.session)

    def initialize_session(self, profile):
        """Initialize AWS session"""
        try:
            print(f"\nInitializing AWS session with profile: {profile!r}")
            self.session = boto3.Session(profile_name=profile)

            # Test the session by making a simple API call
            sts = self.session.client("sts")
            caller_identity = sts.get_caller_identity()
            self.account_id = caller_identity["Account"]

            # Initialize components after session is created
            self.initialize_components()
            return True
        except ClientError as e:
            if "InvalidClientTokenId" in str(e):
                print(f"\nError: Invalid AWS credentials for profile {profile!r}")
            elif "ExpiredToken" in str(e):
                print(f"\nError: AWS credentials for profile {profile!r} have expired")
            else:
                print(f"\nAWS Error: {str(e)}")
            return False
        except Exception as e:
            print(f"\nError initializing session: {str(e)}")
            print("Please verify your AWS credentials and try again")
            return False

    def collect_region_resources(self, region, selected_services=None):
        """Collect resources for a specific region"""
        print(f"\nCollecting resources in {region}...")
        region_resources = []

        # If no specific services selected, use all components
        services_to_scan = selected_services or self.components.keys()

        for service in services_to_scan:
            if service not in self.components:
                print(f"  Warning: {service} is not a valid service component")
                continue

            component = self.components[service]
            try:
                print(f"  Collecting {service} resources in {region}...")
                resources = component.get_resources(region)

                # Skip empty resource lists
                if not resources:
                    print(f"  No {service} resources found in {region}")
                    continue

                # Add resources to the list
                region_resources.extend(resources)

                # Update resource counts
                count = len(resources)
                print(f"  Found {count} {service} resources in {region}")

                # Thread-safe update of instance counts
                with self.resources_lock:
                    if service not in self.instance_counts:
                        self.instance_counts[service] = {}
                    if region not in self.instance_counts[service]:
                        self.instance_counts[service][region] = 0
                    self.instance_counts[service][region] = count

            except Exception as e:
                print(f"  Error collecting {service} resources in {region}: {str(e)}")
                continue

        # Thread-safe update of resources dictionary
        with self.resources_lock:
            self.resources[region] = region_resources

        return region

    def collect_resources(self, selected_services=None):
        """Collect resources across selected regions using parallel processing"""
        if not self.session:
            print(
                "AWS session not initialized. Please call initialize_session() first."
            )
            return False

        print("\nCollecting resources from regions:")
        for region in self.regions:
            print(f"- {region}")

        if selected_services:
            print("\nScanning only these services:")
            for service in selected_services:
                print(f"- {service}")

        # Use ThreadPoolExecutor with max_workers=2 for parallel processing
        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
            # Submit all regions for processing
            future_to_region = {
                executor.submit(
                    self.collect_region_resources, region, selected_services
                ): region
                for region in self.regions
            }

            # Process completed futures
            for future in concurrent.futures.as_completed(future_to_region):
                region = future_to_region[future]
                try:
                    future.result()
                except Exception as e:
                    print(f"Error processing region {region}: {str(e)}")

    def export_to_excel(self, selected_services=None):
        """Export collected resources to Excel"""
        if not self.resources:
            print("No resources collected. Please call collect_resources() first.")
            return False

        try:
            # Create a new workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create summary sheet
            summary_ws = wb.create_sheet("Summary")

            # Style definitions
            header_fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True, size=20)
            regular_font = Font(size=20)
            not_found_fill = PatternFill(
                start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"
            )
            not_found_font = Font(color="FF0000", size=20)

            # Critical fields that should always appear first
            critical_fields = [
                "Region",
                "Service",
                "Resource Name",
                "Resource ID",
                "Description",
                "Creation Time",
            ]

            # Group resources by service
            service_resources = {}
            for region_resources in self.resources.values():
                for resource in region_resources:
                    service = resource["Service"]
                    if service not in service_resources:
                        service_resources[service] = []
                    service_resources[service].append(resource)

            # Create detail sheets
            for service, resources in service_resources.items():
                if not resources:
                    continue

                # Create a sheet name that's valid for Excel
                sheet_name = service.replace(" ", "_")[
                    :31
                ]  # Excel limits sheet names to 31 chars
                ws = wb.create_sheet(sheet_name)

                # Get all headers from data
                all_headers = set()
                for resource in resources:
                    all_headers.update(resource.keys())

                # Sort headers with critical fields first
                headers = critical_fields + sorted(all_headers - set(critical_fields))

                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font

                # Write data
                for row, resource in enumerate(resources, 2):
                    for col, header in enumerate(headers, 1):
                        value = resource.get(header, "")
                        # Format the value based on the header
                        if header == "Creation Time" and value:
                            # Try to format the timestamp if it's not empty
                            try:
                                if isinstance(value, str):
                                    # Remove any timezone info and milliseconds for Excel compatibility
                                    value = value.split("+")[0].split(".")[0]
                            except:
                                pass
                        elif header in ["Tags", "Security Groups", "Subnets"]:
                            # Format list-like data for better readability
                            if isinstance(value, (list, set)):
                                value = "\n".join(sorted(value))
                            elif isinstance(value, dict):
                                value = "\n".join(
                                    f"{k}={v}" for k, v in sorted(value.items())
                                )

                        cell = ws.cell(row=row, column=col, value=str(value))
                        cell.font = regular_font

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min((max_length + 2), 50)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

            # Prepare summary data
            summary_headers = ["Service", "Region", "Resource Count"]
            for col, header in enumerate(summary_headers, 1):
                cell = summary_ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Write summary data for selected services or all services
            services_to_summarize = (
                selected_services if selected_services else self.supported_services
            )
            row = 2
            for service in services_to_summarize:
                for region in self.regions:
                    # Count resources for this service and region
                    count = 0
                    for region_resources in self.resources.values():
                        # Count resources matching this service and region
                        matching_resources = [
                            r
                            for r in region_resources
                            if (
                                r["Service"] == service
                                or (
                                    service == "APIGateway"
                                    and "API Gateway" in r["Service"]
                                )
                                or (service == "Gateway" and r["Service"] == "Gateway")
                                or (
                                    service == "CloudFront"
                                    and r["Service"] == "CloudFront"
                                )
                            )
                            and r["Region"] == region
                        ]
                        count += len(matching_resources)

                    # Write row with count or "Not Found"
                    service_cell = summary_ws.cell(row=row, column=1, value=service)
                    region_cell = summary_ws.cell(row=row, column=2, value=region)
                    count_cell = summary_ws.cell(
                        row=row, column=3, value=count if count > 0 else "Not Found"
                    )

                    # Apply styling
                    service_cell.font = regular_font
                    region_cell.font = regular_font
                    if count == 0:
                        service_cell.fill = not_found_fill
                        region_cell.fill = not_found_fill
                        count_cell.fill = not_found_fill
                        service_cell.font = not_found_font
                        region_cell.font = not_found_font
                        count_cell.font = not_found_font
                    else:
                        count_cell.font = regular_font

                    row += 1

            # Auto-adjust summary column widths
            for column in summary_ws.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min((max_length + 2), 50)
                summary_ws.column_dimensions[column[0].column_letter].width = (
                    adjusted_width
                )

            # Move summary sheet to first position
            wb.move_sheet("Summary", offset=-len(wb.sheetnames))

            # Save the workbook
            file_name = f"aws_inventory_{self.account_id}_{self.timestamp}.xlsx"
            wb.save(file_name)
            print(f"\nInventory saved to {file_name}")
            print("Sheets created:")
            for sheet in wb.sheetnames:
                print(f"- {sheet}")
            return True
        except PermissionError:
            print(
                f"\nError: Unable to save file. The file may be open in another program."
            )
            return False
        except Exception as e:
            print(f"\nError saving Excel file: {str(e)}")
            return False


def main():
    try:
        # Create inventory instance
        inventory = AWSResourceInventory()
        print("\nAWS Resource Inventory Collector")

        # Get AWS profile
        profile = inventory.get_aws_profile()

        # Initialize session with selected profile
        try:
            session = boto3.Session(profile_name=profile)
            sts = session.client("sts")
            account_id = sts.get_caller_identity()["Account"]
            print(f"\nUsing profile {profile!r}")
            inventory.session = session
            inventory.account_id = account_id
            inventory.initialize_components()  # Initialize components after session
        except (
            botocore.exceptions.ProfileNotFound,
            botocore.exceptions.ClientError,
        ) as e:
            print(f"\nProfile {profile!r} not found or invalid")
            sys.exit(1)

        print(f"\nConnected to AWS Account: {account_id}")

        # Select regions
        inventory.select_regions()

        # Show available services and let user select
        print("\nAvailable services:")
        for i, service in enumerate(inventory.supported_services, 1):
            print(f"{i}. {service}")

        print(
            "\nSelect services to scan (comma-separated numbers, or 'all' for all services)"
        )
        selected_services = None
        while True:
            selection = input("\nEnter selection: ").strip().lower()
            if selection == "all":
                break
            try:
                indices = [int(x.strip()) for x in selection.split(",")]
                selected_services = [
                    inventory.supported_services[i - 1]
                    for i in indices
                    if 0 < i <= len(inventory.supported_services)
                ]
                break
            except (ValueError, IndexError):
                print("\nInvalid selection. Please try again.")

        # Collect resources
        print("\nCollecting resources...")
        inventory.collect_resources(selected_services)

        # Export to Excel
        print("\nExporting to Excel...")
        if inventory.export_to_excel(selected_services):
            print("\nExport completed successfully!")
        else:
            print("\nExport failed.")

    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        sys.exit(1)
    except Exception as e:
        print(f"\nUnexpected error: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
